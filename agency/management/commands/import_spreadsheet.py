# agency/management/commands/import_spreadsheet.py
from django.core.management.base import BaseCommand
from django.db import transaction
import openpyxl
from decimal import Decimal
from datetime import date
from ...models import Company, Client, UserProfile, MonthlyRevenue, Project
from django.contrib.auth.models import User

class Command(BaseCommand):
    help = 'Import data from Excel spreadsheet'
    
    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to Excel file')
        parser.add_argument('company_code', type=str, help='Company code to import data for')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Perform a dry run without saving data',
        )
    
    def handle(self, *args, **options):
        file_path = options['file_path']
        company_code = options['company_code']
        dry_run = options['dry_run']
        
        # Get or create company
        company, created = Company.objects.get_or_create(
            code=company_code,
            defaults={'name': f'Company {company_code}'}
        )
        
        if created:
            self.stdout.write(f'Created company: {company.name}')
        else:
            self.stdout.write(f'Using existing company: {company.name}')
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            if dry_run:
                self.stdout.write(self.style.WARNING('DRY RUN - No data will be saved'))
                self.preview_import(workbook)
            else:
                with transaction.atomic():
                    results = self.import_data(workbook, company)
                    self.display_results(results)
                    
        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f'File not found: {file_path}')
            )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Import failed: {str(e)}')
            )
    
    def import_data(self, workbook, company):
        """Import data from workbook"""
        results = {
            'clients_created': 0,
            'revenue_entries': 0,
            'users_created': 0,
            'projects_created': 0,
            'errors': []
        }
        
        # Import revenue data
        if 'Revenue' in workbook.sheetnames:
            try:
                revenue_results = self.import_revenue_sheet(workbook['Revenue'], company)
                results['clients_created'] += revenue_results['clients_created']
                results['revenue_entries'] += revenue_results['revenue_entries']
                results['projects_created'] += revenue_results['projects_created']
            except Exception as e:
                results['errors'].append(f"Revenue import error: {str(e)}")
        
        # Import payroll data
        if 'Payroll' in workbook.sheetnames:
            try:
                payroll_results = self.import_payroll_sheet(workbook['Payroll'], company)
                results['users_created'] += payroll_results['users_created']
            except Exception as e:
                results['errors'].append(f"Payroll import error: {str(e)}")
        
        return results
    
    def import_revenue_sheet(self, sheet, company):
        """Import revenue data from Revenue sheet"""
        results = {'clients_created': 0, 'revenue_entries': 0, 'projects_created': 0}
        
        # Skip header row, iterate through data rows
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
                
            client_name = str(row[0]).strip()
            if not client_name:
                continue
                
            # Get status from column C (index 2)
            status = str(row[2]).strip() if len(row) > 2 and row[2] else 'Open'
            client_status = 'active' if status.lower() == 'open' else 'inactive'
            
            self.stdout.write(f"Processing client: {client_name} (Status: {status})")
            
            # Create or get client
            client, created = Client.objects.get_or_create(
                name=client_name,
                company=company,
                defaults={'status': client_status}
            )
            
            if created:
                results['clients_created'] += 1
                self.stdout.write(f"  Created client: {client_name}")
            
            # Create a general project for this client
            project_name = f"{client_name} - General Work"
            project, project_created = Project.objects.get_or_create(
                name=project_name,
                client=client,
                company=company,
                defaults={
                    'start_date': date(2025, 1, 1),
                    'end_date': date(2025, 12, 31),
                    'total_revenue': Decimal('0'),
                    'total_hours': Decimal('0'),
                    'status': 'active' if client_status == 'active' else 'completed'
                }
            )
            
            if project_created:
                results['projects_created'] += 1
            
            # Import monthly revenue (columns 3-14 for Jan-Dec, but accounting for offset)
            total_revenue = Decimal('0')
            for month in range(1, 13):
                col_index = month + 2  # Columns D through O (3-14)
                if col_index < len(row) and row[col_index]:
                    try:
                        revenue_amount = Decimal(str(row[col_index]))
                        if revenue_amount > 0:
                            total_revenue += revenue_amount
                            
                            MonthlyRevenue.objects.update_or_create(
                                client=client,
                                company=company,
                                year=2025,
                                month=month,
                                revenue_type='booked',
                                defaults={'revenue': revenue_amount}
                            )
                            results['revenue_entries'] += 1
                            
                    except (ValueError, TypeError, Exception) as e:
                        self.stdout.write(f"  Warning: Could not parse revenue for {client_name}, month {month}: {e}")
                        continue
            
            # Update project total revenue
            if total_revenue > 0:
                project.total_revenue = total_revenue
                project.save()
                self.stdout.write(f"  Total revenue for {client_name}: ${total_revenue}")
        
        return results
    
    def import_payroll_sheet(self, sheet, company):
        """Import user data from Payroll sheet"""
        results = {'users_created': 0}
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            full_name = str(row[0]).strip()
            if not full_name:
                continue
                
            self.stdout.write(f"Processing team member: {full_name}")
            
            # Parse name
            name_parts = full_name.split()
            first_name = name_parts[0] if name_parts else 'Unknown'
            last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
            
            # Create username from name
            username = f"{first_name.lower()}.{last_name.lower()}".replace(' ', '.').replace('-', '.')
            
            # Get salary info
            annual_salary = Decimal('0')
            if len(row) > 1 and row[1]:
                try:
                    annual_salary = Decimal(str(row[1]))
                except (ValueError, TypeError):
                    annual_salary = Decimal('0')
            
            # Calculate hourly rate (assuming 2080 hours per year)
            hourly_rate = annual_salary / 2080 if annual_salary > 0 else Decimal('75.00')
            
            # Create user if doesn't exist
            user, user_created = User.objects.get_or_create(
                username=username,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': f"{username}@{company.code.lower()}.com",
                }
            )
            
            if user_created:
                self.stdout.write(f"  Created user: {username}")
            
            # Create or update user profile
            profile, profile_created = UserProfile.objects.get_or_create(
                user=user,
                defaults={
                    'company': company,
                    'role': 'tech',  # Default role
                    'hourly_rate': hourly_rate,
                    'annual_salary': annual_salary if annual_salary > 0 else None,
                    'status': 'full_time',
                    'start_date': date(2025, 1, 1),
                    'weekly_capacity_hours': Decimal('40'),
                    'utilization_target': Decimal('80')
                }
            )
            
            if profile_created:
                results['users_created'] += 1
                self.stdout.write(f"  Created profile for: {full_name} (${hourly_rate}/hr)")
            elif not profile_created:
                # Update existing profile
                profile.hourly_rate = hourly_rate
                profile.annual_salary = annual_salary if annual_salary > 0 else None
                profile.save()
                self.stdout.write(f"  Updated profile for: {full_name}")
        
        return results
    
    def preview_import(self, workbook):
        """Preview what would be imported"""
        self.stdout.write(self.style.SUCCESS('IMPORT PREVIEW:'))
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Count non-empty rows
            row_count = 0
            for row in sheet.iter_rows(min_row=2):
                if any(cell.value for cell in row):
                    row_count += 1
            
            self.stdout.write(f'  {sheet_name} sheet: {row_count} data rows')
            
            # Show first few entries
            if sheet_name == 'Revenue':
                self.stdout.write(f'    Sample clients:')
                count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and count < 3:
                        self.stdout.write(f'      - {row[0]}')
                        count += 1
            
            elif sheet_name == 'Payroll':
                self.stdout.write(f'    Sample team members:')
                count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and count < 3:
                        self.stdout.write(f'      - {row[0]}')
                        count += 1
    
    def display_results(self, results):
        """Display import results"""
        self.stdout.write(self.style.SUCCESS('IMPORT COMPLETED:'))
        self.stdout.write(f'  Clients created: {results["clients_created"]}')
        self.stdout.write(f'  Projects created: {results["projects_created"]}')
        self.stdout.write(f'  Users created: {results["users_created"]}')
        self.stdout.write(f'  Revenue entries: {results["revenue_entries"]}')
        
        if results['errors']:
            self.stdout.write(self.style.ERROR('ERRORS:'))
            for error in results['errors']:
                self.stdout.write(f'  - {error}')
